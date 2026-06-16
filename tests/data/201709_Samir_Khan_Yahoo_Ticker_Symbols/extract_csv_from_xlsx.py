from openpyxl import load_workbook
import csv
from pathlib import Path

outdir = Path('tests/data/201709_Samir_Khan_Yahoo_Ticker_Symbols')

wb = load_workbook('Yahoo Ticker Symbols - September 2017.xlsx', read_only=True)
ss = wb.sheetnames[:-1]
zz = set()
def process_sheet(s):
    print(f"==== {s}")
    rows = []
    empty = set(['simulationconsultant@gmail.com','Samir Khan'])

    def is_empty(*vv):
        return all(v is None or v in empty for v in vv)

    def find_empty_tail(row)->int:
        for i, v in enumerate(row):
            if is_empty(v):
                break
        return i
    
    for i,row in enumerate(wb[s].iter_rows(values_only=True)):
        if i < 3:
            continue
        if not is_empty(*row):
            rows.append(row)

    p = find_empty_tail(rows[0])
    h = rows[0][:p]
    if 'Exchange' not in h:
        h = [*h, 'Exchange']
        p += 1 
    data = [row[:p] for row in rows[1:]]
    not_empty = set([tuple(row[p:]) for row in rows[1:] if not is_empty(*row[p:])])
    print(h)
    for i in range(3):
        print(data[i])
    print (f'... total rows:{len(data)}')
    for i in range(3):
        print(data[-3+i])
    not_empty.discard(tuple())
    if not_empty:
        print(not_empty)
        for ne in not_empty:
            zz.add(ne[0])
    else:
        f = outdir / f'{s.replace(' ', '_')}.csv'
        with open(f, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows([h, *data])
        print(f"Writen: {f}")

    print('---')
  
  
for s in ss:
  process_sheet(s)

print('===')
print(zz)

